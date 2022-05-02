import numpy as np
import pandas as pd
from pandas_datareader import data as wb
from openpyxl import Workbook
from openpyxl import load_workbook

#workbook = Workbook()
workbook = load_workbook(filename="Test.xlsx", keep_vba=True, data_only=True)

sheet = workbook.active

sheet["B5"] = 3

workbook.save(filename="Test.xlsx")
